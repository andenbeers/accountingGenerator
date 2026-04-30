#!/usr/bin/env python3
"""
PDF → docstrange (extract_data json_schema) → Excel
"""

import json
import os
import math
from collections import defaultdict

os.environ["PATH"] = os.path.abspath(r"poppler\library\bin") + ";" + os.environ["PATH"]

from docstrange import DocumentExtractor
from pypdf import PdfReader, PdfWriter
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ── CONFIG ────────────────────────────────────────────────────────────────────

PDF_PATH   = "input4.pdf"
PAGE_START = 4
PAGE_END   = 6
OUTPUT_XLS = "transactions.xlsx"

extractor = DocumentExtractor(gpu=False)

# ── SCHEMA ────────────────────────────────────────────────────────────────────

SCHEMA = {
    "purchases": [
        {
            "date":        "string",
            "cusip":       "string",
            "description": "string",
            "quantity":    "number",
            "amount":      "number",
        }
    ],
    "sales": [
        {
            "date":               "string",
            "cusip":              "string",
            "description":        "string",
            "quantity":           "number",
            "amount":             "number",
            "realized_gain_loss": "number",
            "carry_value":        "number",
        }
    ],
    "dividends": [
        {
            "date":        "string",
            "cusip":       "string",
            "description": "string",
            "amount":      "number",
            "action":      "string",
        }
    ],
    # Only extract interest rows from the Transaction Details table
    # where Category == "Interest". Bank Sweep Activity entries are NOT interest.
    "interest": [
        {
            "date":        "string",
            "description": "string",
            "amount":      "number",
        }
    ],
}


# ── STEP 1: PDF → STRUCTURED DATA ────────────────────────────────────────────

def pdf_page_to_data(pdf_path: str, page_idx: int) -> dict:
    writer = PdfWriter()
    writer.add_page(PdfReader(pdf_path).pages[page_idx])
    tmp = f"/tmp/_page_{page_idx + 1}.pdf"
    with open(tmp, "wb") as f:
        writer.write(f)
    result = extractor.extract(tmp)
    data = result.extract_data(json_schema=SCHEMA)

    # unwrap docstrange nesting: {"structured_data": {"content": {...}}}
    if isinstance(data, dict):
        if "structured_data" in data:
            data = data["structured_data"]
        if "content" in data:
            data = data["content"]

    return data if isinstance(data, dict) else {}


def extract_pages(pdf_path: str, start: int, end: int) -> list[tuple[int, dict]]:
    pages = []
    for idx in range(start - 1, end):
        print(f"  docstrange → page {idx + 1} …", end=" ", flush=True)
        data = pdf_page_to_data(pdf_path, idx)
        print(f"done")
        print(f"  preview: {json.dumps(data)[:300]}\n")
        if data:
            pages.append((idx + 1, data))
        else:
            print(f"  [warn] page {idx + 1} returned no data, skipping")
    return pages


# ── STEP 2: MERGE + DEDUPLICATE ───────────────────────────────────────────────

def merge(base: dict, new: dict) -> None:
    for key in base:
        items = new.get(key, [])
        if isinstance(items, list):
            base[key].extend(items)


def _safe_amount(item: dict) -> float:
    """Parse amount robustly regardless of type."""
    val = item.get("amount") or 0
    try:
        return round(float(str(val).replace(",", "").strip()), 2)
    except (ValueError, TypeError):
        return 0.0


def _interest_key(item: dict) -> tuple:
    """
    Dedup key for interest: (rounded_amount, month).
    The same payment appears in Transaction Details (date=09/16)
    and Bank Sweep Activity (date=09/15) — same month, same amount.
    Keeping only the first occurrence (Transaction Details) is correct.
    """
    amt = _safe_amount(item)
    date = str(item.get("date") or "")
    # Use just the month portion so 09/15 and 09/16 collapse together
    month = date.split("/")[0] if "/" in date else date[:7]
    return (amt, month)


def deduplicate(data: dict) -> dict:
    # Generic dedup by full JSON equality for all sections
    for section, items in data.items():
        seen, out = set(), []
        for item in items:
            k = json.dumps(item, sort_keys=True)
            if k not in seen:
                seen.add(k)
                out.append(item)
        removed = len(items) - len(out)
        if removed:
            print(f"  [fix] removed {removed} exact duplicate(s) from '{section}'")
        data[section] = out

    # Interest-specific filters
    clean = []
    for item in data["interest"]:
        date = str(item.get("date") or "")
        desc = str(item.get("description") or "").lower()

        # Drop summary rows with no day (e.g. "2024-09")
        if date.count("-") >= 1 and date.count("-") < 2 and len(date) <= 7:
            print(f"  [fix] dropped interest — no day in date: {date!r}")
            continue

        # Drop rate rows
        if "rate" in desc:
            print(f"  [fix] dropped interest — looks like a rate: {desc!r}")
            continue

        # Drop Bank Sweep Activity echo rows — they say "bank interest" or
        # "bank sweep" in the description but come from the sweep section
        if "bank sweep" in desc or ("bank interest" in desc and "bank int " not in desc):
            print(f"  [fix] dropped interest — bank sweep echo: {desc!r}")
            continue

        clean.append(item)

    # Dedup by (amount, month) — catches same payment with different day
    seen_key, deduped = set(), []
    for item in clean:
        k = _interest_key(item)
        if k not in seen_key:
            seen_key.add(k)
            deduped.append(item)
        else:
            print(f"  [fix] dropped interest duplicate: amount={k[0]} month={k[1]}")

    removed = len(data["interest"]) - len(deduped)
    if removed:
        print(f"  [fix] interest: kept {len(deduped)} of {len(data['interest'])} entries")
    data["interest"] = deduped
    return data


# ── STEP 3: EXCEL OUTPUT ──────────────────────────────────────────────────────

MONEY = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

class Sheet:
    def __init__(self):
        self.wb  = openpyxl.Workbook()
        self.ws  = self.wb.active
        self.ws.title = "Transactions"
        self.row = 1

    def write(self, cells, money_cols=(), bold=False):
        money_cols = set(money_cols)
        for col, val in enumerate(cells, 1):
            c = self.ws.cell(row=self.row, column=col, value=val)
            is_money = isinstance(val, (int, float)) or (isinstance(val, str) and val.startswith('='))
            if (col - 1) in money_cols and is_money:
                c.number_format = MONEY
            if bold:
                c.font = Font(bold=True)
        self.row += 1

    def blank(self, n=1):
        self.row += n

    def sum_col(self, col: int, r1: int, r2: int) -> str:
        return f"=SUM({get_column_letter(col)}{r1}:{get_column_letter(col)}{r2})"

    def save(self, path: str):
        self.wb.save(path)
        print(f"Saved → {path}")

    def purchases(self, items: list):
        if not items:
            return
        self.write(["Purchases"], bold=True)
        for name, rows in _group(items, "description").items():
            self.blank()
            self.write([name], bold=True)
            self.write(["Date", "Quantity", "Amount"], bold=True)
            r0 = self.row
            for r in rows:
                self.write([r.get("date"), _shares(r), _abs(r, "amount")], money_cols=[2])
            r1 = self.row - 1
            self.write(["", "Total", self.sum_col(3, r0, r1)], money_cols=[2], bold=True)
        self.blank(2)

    def sales(self, items: list):
        if not items:
            return
        self.write(["Sales"], bold=True)
        for name, rows in _group(items, "description").items():
            self.blank()
            self.write([name], bold=True)
            self.write(["Date", "Quantity", "Carry Value", "Sale Amount", "Gain", "Loss"], bold=True)
            r0 = self.row
            for r in rows:
                gl  = float(r.get("realized_gain_loss") or 0)
                amt = float(r.get("amount") or 0)
                cv  = float(r.get("carry_value") or 0) or (amt - gl)
                self.write(
                    [r.get("date"), _shares(r), cv, amt,
                     gl if gl > 0 else 0, -gl if gl < 0 else 0],
                    money_cols=[2, 3, 4, 5],
                )
            r1 = self.row - 1
            self.write(
                ["", "Total",
                 self.sum_col(3, r0, r1), self.sum_col(4, r0, r1),
                 self.sum_col(5, r0, r1), self.sum_col(6, r0, r1)],
                money_cols=[2, 3, 4, 5], bold=True,
            )
        self.blank(2)

    def dividends(self, items: list):
        if not items:
            return
        self.write(["Dividends"], bold=True)
        self.write(["Date", "Description", "Amount"], bold=True)
        r0 = self.row
        for r in items:
            self.write([r.get("date"), r.get("description"), _abs(r, "amount")], money_cols=[2])
        r1 = self.row - 1
        self.write(["", "Total", self.sum_col(3, r0, r1)], money_cols=[2], bold=True)
        self.blank(2)

    def interest(self, items: list):
        if not items:
            return
        self.write(["Interest"], bold=True)
        self.write(["Date", "Description", "Amount"], bold=True)
        r0 = self.row
        for r in items:
            self.write([r.get("date"), r.get("description", ""), _abs(r, "amount")], money_cols=[2])
        r1 = self.row - 1
        self.write(["", "Total", self.sum_col(3, r0, r1)], money_cols=[2], bold=True)


# ── HELPERS ───────────────────────────────────────────────────────────────────

_ticker_cache: dict[str, str] = {}

def _resolve_name(cusip: str | None, fallback: str | None) -> str:
    symbol = (cusip or "").strip().upper()
    if symbol and symbol not in _ticker_cache:
        try:
            import yfinance as yf
            info = yf.Ticker(symbol).info
            name = (info.get("shortName") or info.get("longName") or "").strip()
            _ticker_cache[symbol] = name or _title(fallback or symbol)
            print(f"  [ticker] {symbol} → {_ticker_cache[symbol]}")
        except Exception:
            _ticker_cache[symbol] = _title(fallback or symbol)
    return _ticker_cache.get(symbol) or _title(fallback or "Unknown")

def _title(s: str) -> str:
    return s.strip().title() if s else s

def resolve_names(data: dict) -> dict:
    for section in data.values():
        for item in section:
            item["description"] = _resolve_name(item.get("cusip"), item.get("description"))
    return data

def _group(items: list, key: str) -> dict:
    out = defaultdict(list)
    for item in items:
        out[item.get(key) or "Unknown"].append(item)
    return out

def _abs(item: dict, key: str) -> float:
    return math.fabs(float(item.get(key) or 0))

def _shares(item: dict) -> str:
    q = item.get("quantity")
    return f"{abs(float(q)):.3f} shares" if q else ""


# ── MAIN ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"\n[1] Extracting pages {PAGE_START}–{PAGE_END} with docstrange …")
    pages = extract_pages(PDF_PATH, PAGE_START, PAGE_END)

    print("\n[2] Merging pages …")
    data: dict = {"dividends": [], "purchases": [], "sales": [], "interest": []}
    for page_num, raw in pages:
        print(f"  page {page_num} …")
        merge(data, raw)

    print("\n[3] Deduplicating …")
    data = deduplicate(data)
    print(json.dumps(data, indent=2))

    print("\n[4] Resolving ticker names …")
    data = resolve_names(data)

    print("\n[5] Writing Excel …")
    sheet = Sheet()
    sheet.purchases(data["purchases"])
    sheet.sales(data["sales"])
    sheet.dividends(data["dividends"])
    sheet.interest(data["interest"])
    sheet.save(OUTPUT_XLS)